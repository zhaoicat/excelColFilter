#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
命令行Excel列导出工具
支持通过命令行参数指定输入文件、输出文件和要导出的列
"""

import argparse
import pandas as pd
import sys
import os
import requests
import time
from openpyxl.drawing import image as xl_image
import hashlib
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


def read_excel_file(file_path):
    """读取Excel文件"""
    try:
        if file_path.endswith('.xls'):
            df = pd.read_html(file_path, encoding='utf-8', header=0)[0]
        else:
            df = pd.read_excel(file_path, dtype=str)
        
        # 将可能包含长数字的列转换为字符串，避免科学计数法
        id_columns = ['主订单号', '子订单号', '店铺ID', '商品ID', '规格编号', 
                     '采购订单号', '平台物流单号', '手机号', '商户订单号']
        
        for col in id_columns:
            if col in df.columns:
                # 确保转换为字符串并处理科学计数法
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(df[col])
                df[col] = df[col].apply(lambda x: f"{x:.0f}" if pd.notna(x) and isinstance(x, (int, float)) else str(x))
                # 移除可能的 .0 后缀和 nan 值
                df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)
                df[col] = df[col].replace('nan', '')
        
        print(f"成功读取文件: {file_path}")
        print(f"数据形状: {df.shape[0]} 行 x {df.shape[1]} 列")
        return df
    
    except Exception as e:
        print(f"读取文件失败: {e}")
        return None


def get_columns_from_input(input_string, available_columns):
    """解析用户输入的列信息"""
    selected_columns = []
    
    if input_string.lower() == 'all':
        return available_columns
    
    # 首先尝试作为列名处理
    column_names = [name.strip() for name in input_string.split(',')]
    
    # 检查是否所有部分都是纯数字或数字范围
    is_numeric_selection = True
    for part in column_names:
        part = part.strip()
        if '-' in part:
            # 检查是否是数字范围
            try:
                start, end = part.split('-')
                int(start.strip())
                int(end.strip())
            except (ValueError, AttributeError):
                is_numeric_selection = False
                break
        else:
            # 检查是否是单个数字
            try:
                int(part)
            except ValueError:
                is_numeric_selection = False
                break
    
    if is_numeric_selection:
        # 处理数字选择
        for part in column_names:
            part = part.strip()
            if '-' in part:
                # 处理范围 (例如: 1-5)
                try:
                    start, end = map(int, part.split('-'))
                    for i in range(start, end + 1):
                        if 1 <= i <= len(available_columns):
                            selected_columns.append(available_columns[i-1])
                except ValueError:
                    print(f"无效的范围格式: {part}")
            else:
                # 处理单个数字
                try:
                    num = int(part)
                    if 1 <= num <= len(available_columns):
                        selected_columns.append(available_columns[num-1])
                    else:
                        print(f"列序号超出范围: {num}")
                except ValueError:
                    print(f"无效的数字: {part}")
    else:
        # 处理列名选择
        for name in column_names:
            if name in available_columns:
                selected_columns.append(name)
            else:
                print(f"警告: 列 '{name}' 不存在")
    
    # 去重但保持顺序
    seen = set()
    result = []
    for col in selected_columns:
        if col not in seen:
            seen.add(col)
            result.append(col)
    return result


def convert_to_bmp_in_memory(image_path):
    """在内存中将任意格式图片转换为BMP格式的临时文件用于Excel显示"""
    if not PIL_AVAILABLE:
        return image_path  # 如果PIL不可用，返回原路径
        
    try:
        import tempfile
        
        with Image.open(image_path) as img:
            # 转换为RGB模式（BMP不支持透明通道）
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                if img.mode == 'RGBA':
                    background.paste(img, mask=img.split()[-1])
                else:
                    background.paste(img)
                img = background
            else:
                img = img.convert('RGB')
            
            # 创建临时BMP文件
            temp_file = tempfile.NamedTemporaryFile(suffix='.bmp', delete=False)
            temp_path = temp_file.name
            temp_file.close()
            
            # 保存为BMP格式到临时文件
            img.save(temp_path, 'BMP')
            
            return temp_path
    except Exception as e:
        print(f"图片内存转换BMP失败 {image_path}: {e}")
        return image_path  # 转换失败，返回原路径


def download_single_image(url, output_dir="images", timeout=30, max_retries=3):
    """下载单个图片（保持原图），支持重试"""
    # 验证URL
    if not url or pd.isna(url) or not str(url).startswith(('http://', 'https://')):
        return None, url
    
    # 创建图片目录
    os.makedirs(output_dir, exist_ok=True)
    
    # 生成文件名（使用URL的hash值避免重复和特殊字符）
    url_hash = hashlib.md5(str(url).encode()).hexdigest()
    
    # 尝试从URL获取文件扩展名
    parsed_url = urlparse(str(url))
    path = parsed_url.path.lower()
    if path.endswith(('.jpg', '.jpeg')):
        ext = '.jpg'
    elif path.endswith('.png'):
        ext = '.png'
    elif path.endswith('.gif'):
        ext = '.gif'
    elif path.endswith('.webp'):
        ext = '.webp'
    else:
        ext = '.jpg'  # 默认使用jpg
    
    filename = f"{url_hash}{ext}"
    filepath = os.path.join(output_dir, filename)
    
    # 如果文件已存在，直接返回
    if os.path.exists(filepath):
        return filepath, url
    
    # 重试下载
    last_error = None
    for attempt in range(max_retries):
        try:
            # 下载图片
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(str(url), headers=headers, timeout=timeout, stream=True)
            response.raise_for_status()
            
            # 保存图片
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            # 验证文件是否成功保存
            if os.path.exists(filepath):
                return filepath, url
            else:
                raise Exception("文件保存后不存在")
                
        except Exception as e:
            last_error = e
            if attempt < max_retries - 1:  # 如果不是最后一次尝试
                print(f"下载失败，第{attempt + 1}次重试: {url}")
                time.sleep(1)  # 等待1秒后重试
            continue
    
    print(f"下载最终失败 (重试{max_retries}次): {url}, 错误: {last_error}")
    return None, url


def download_images_parallel(urls, max_workers=5, output_dir="images"):
    """并行下载多个图片"""
    results = {}
    failed_urls = []
    
    print(f"\n开始并行下载 {len(urls)} 张图片（最多 {max_workers} 个线程）...")
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有下载任务
        future_to_url = {
            executor.submit(download_single_image, url, output_dir): url 
            for url in urls if pd.notna(url) and str(url).strip()
        }
        
        completed = 0
        total = len(future_to_url)
        
        # 处理完成的任务
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                filepath, original_url = future.result()
                if filepath and os.path.exists(filepath):
                    results[original_url] = filepath
                else:
                    failed_urls.append(original_url)
                    if filepath:
                        print(f"下载的文件不存在: {filepath}")
            except Exception as e:
                print(f"下载异常 {url}: {e}")
                failed_urls.append(url)
            
            completed += 1
            print(f"下载进度: {completed}/{total} (成功: {len(results)}, 失败: {len(failed_urls)})", end='\r')
    
    print(f"\n下载完成！成功: {len(results)}, 失败: {len(failed_urls)}")
    
    if failed_urls:
        print("失败的URL示例:", failed_urls[:3])
    
    return results


def export_columns(df, selected_columns, output_file, download_images=False):
    """导出指定列到Excel文件"""
    try:
        if not selected_columns:
            print("错误: 没有有效的列可以导出")
            return False
        
        # 检查是否有商品图片列且需要下载图片
        image_column = None
        if download_images:
            if '商品图片' in selected_columns:
                image_column = '商品图片'
        
        # 如果有图片列且要下载图片，调整列顺序，将图片列放在第一列
        if download_images and image_column:
            # 将图片列移到第一列
            reordered_columns = [image_column]
            for col in selected_columns:
                if col != image_column:
                    reordered_columns.append(col)
            selected_columns = reordered_columns
        
        result_df = df[selected_columns].copy()
        
        # 使用ExcelWriter来控制格式
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # 获取工作表
            worksheet = writer.sheets['Sheet1']
            
            # 将可能的ID列设置为文本格式
            id_columns = ['主订单号', '子订单号', '店铺ID', '商品ID', '规格编号', 
                         '采购订单号', '平台物流单号', '手机号', '商户订单号']
            
            # 找到ID列的位置并设置为文本格式
            for col_idx, col_name in enumerate(result_df.columns, 1):
                if col_name in id_columns:
                    for row_idx in range(2, len(result_df) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = '@'  # 文本格式
            
            # 如果需要下载图片且找到了图片列
            if download_images and image_column:
                # 收集所有图片URL
                image_urls = []
                empty_urls = 0
                duplicate_urls = set()
                for _, row in result_df.iterrows():
                    url = row[image_column]
                    if pd.notna(url) and str(url).strip():
                        url_str = str(url)
                        if url_str in duplicate_urls:
                            continue  # 跳过重复URL
                        duplicate_urls.add(url_str)
                        image_urls.append(url_str)
                    else:
                        empty_urls += 1
                
                print(f"URL统计: 总行数={len(result_df)}, 有效URL={len(image_urls)}, 空URL={empty_urls}, 重复URL={len(result_df)-len(image_urls)-empty_urls}")
                
                temp_files_to_cleanup = []  # 初始化临时文件列表
                if image_urls:
                    # 并行下载所有图片
                    image_results = download_images_parallel(image_urls, max_workers=8)
                    
                    # 找到图片列的位置（现在应该在第一列）
                    image_col_idx = None
                    for col_idx, col_name in enumerate(result_df.columns, 1):
                        if col_name == image_column:
                            image_col_idx = col_idx
                            break
                    
                    if image_col_idx:
                        # 设置图片列宽
                        col_letter = worksheet.cell(row=1, column=image_col_idx).column_letter
                        worksheet.column_dimensions[col_letter].width = 15
                        
                        inserted_count = 0
                        
                        # 插入图片到Excel
                        for row_idx, (_, row) in enumerate(result_df.iterrows(), 2):
                            image_url = str(row[image_column]) if pd.notna(row[image_column]) else ""
                            
                            if image_url in image_results:
                                image_path = image_results[image_url]
                                
                                try:
                                    # 确保文件存在
                                    if not os.path.exists(image_path):
                                        print(f"文件不存在，跳过: {image_path}")
                                        continue
                                    
                                    # 统一转换图片格式为BMP以兼容Excel
                                    final_image_path = convert_to_bmp_in_memory(image_path)
                                    if not final_image_path:
                                        print(f"图片格式转换失败，跳过: {image_path}")
                                        continue
                                    
                                    # 如果是新生成的临时文件，记录以便后续清理
                                    if final_image_path != image_path:
                                        temp_files_to_cleanup.append(final_image_path)
                                    
                                    # 插入图片到Excel
                                    img = xl_image.Image(final_image_path)
                                    
                                    # 获取原图尺寸并保持比例缩放显示
                                    original_width = img.width
                                    original_height = img.height
                                    max_display_size = 100
                                    
                                    if original_width > max_display_size or original_height > max_display_size:
                                        if original_width > original_height:
                                            img.width = max_display_size
                                            img.height = int(original_height * max_display_size / original_width)
                                        else:
                                            img.height = max_display_size
                                            img.width = int(original_width * max_display_size / original_height)
                                    
                                    # 设置图片位置
                                    cell_ref = worksheet.cell(row=row_idx, column=image_col_idx).coordinate
                                    img.anchor = cell_ref
                                    
                                    worksheet.add_image(img)
                                    
                                    # 清空单元格中的URL文字，只保留图片
                                    cell = worksheet.cell(row=row_idx, column=image_col_idx)
                                    cell.value = ""
                                    
                                    # 设置行高以适应图片
                                    worksheet.row_dimensions[row_idx].height = 80
                                    
                                    inserted_count += 1
                                    
                                except Exception as e:
                                    print(f"插入图片失败 {image_url}: {e}")
                        
                        print(f"成功插入 {inserted_count} 张图片到Excel")
                else:
                    print("没有找到有效的图片URL")
        
        print("\n导出成功!")
        print(f"导出文件: {output_file}")
        print(f"导出列数: {len(selected_columns)}")
        print(f"数据行数: {result_df.shape[0]}")
        print(f"导出的列: {selected_columns}")
        
        if download_images and image_column:
            print(f"图片处理: 从 {image_column} 列处理了图片并保存到 images/ 目录")
        
        # Excel保存完成后清理临时文件
        if 'temp_files_to_cleanup' in locals():
            for temp_file in temp_files_to_cleanup:
                try:
                    os.remove(temp_file)
                except Exception:
                    pass  # 忽略清理失败
        
        return True
    
    except Exception as e:
        import traceback
        print(f"导出失败: {e}")
        print(f"详细错误信息:\n{traceback.format_exc()}")
        return False


def list_columns(df):
    """显示所有可用的列"""
    print("\n可用的列:")
    print("=" * 60)
    for i, col in enumerate(df.columns, 1):
        print(f"{i:2d}. {col}")
    print("=" * 60)


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='Excel列导出工具 - 从Excel文件中导出指定的列',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
使用示例:
  # 基本用法 - 按列序号选择
  python cli_excel_processor.py -i input.xls -o output.xlsx -c "1,2,5,10-15"
  
  # 按列名选择
  python cli_excel_processor.py -i input.xls -o output.xlsx -c "编号,平台,站点,店铺名称"
  
  # 导出所有列
  python cli_excel_processor.py -i input.xls -o output.xlsx -c "all"
  
  # 仅查看列信息，不导出
  python cli_excel_processor.py -i input.xls --list-columns
  
  # 导出并下载商品图片
  python cli_excel_processor.py -i input.xls -o output.xlsx -c "编号,商品图片,商品标题" --download-images
        '''
    )
    
    parser.add_argument('-i', '--input', 
                       help='输入Excel文件路径')
    
    parser.add_argument('-o', '--output',
                       help='输出Excel文件路径 (默认: output.xlsx)')
    
    parser.add_argument('-c', '--columns',
                       help='要导出的列。支持格式: 列序号(1,2,5,10-15) | 列名(编号,平台,站点) | all(所有列)')
    
    parser.add_argument('--list-columns',
                       action='store_true',
                       help='仅显示所有可用的列，不进行导出')
    
    parser.add_argument('--download-images',
                       action='store_true',
                       help='下载商品图片并在Excel中展示')
    
    args = parser.parse_args()
    
    # 如果没有提供输入文件，显示帮助信息并退出
    if not args.input:
        parser.print_help()
        sys.exit(0)
    
    # 检查输入文件是否存在
    if not os.path.exists(args.input):
        print(f"错误: 输入文件 '{args.input}' 不存在")
        sys.exit(1)
    
    # 读取Excel文件
    df = read_excel_file(args.input)
    if df is None:
        sys.exit(1)
    
    # 如果只是要查看列信息
    if args.list_columns:
        list_columns(df)
        return
    
    # 检查是否提供了列信息
    if not args.columns:
        print("错误: 请使用 -c 参数指定要导出的列，或使用 --list-columns 查看所有可用列")
        list_columns(df)
        sys.exit(1)
    
    # 解析列选择
    available_columns = list(df.columns)
    selected_columns = get_columns_from_input(args.columns, available_columns)
    
    if not selected_columns:
        print("错误: 没有找到有效的列")
        sys.exit(1)
    
    # 设置输出文件名
    output_file = args.output or 'output.xlsx'
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    # 导出数据
    success = export_columns(df, selected_columns, output_file, args.download_images)
    
    if success:
        print(f"\n✅ 处理完成! 文件已保存为: {output_file}")
    else:
        print("\n❌ 处理失败")
        sys.exit(1)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n程序已退出")
        sys.exit(1)
    except Exception as e:
        print(f"\n程序出错: {e}")
        sys.exit(1) 